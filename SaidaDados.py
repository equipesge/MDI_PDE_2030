from Sistema import Sistema;
from Problema import Problema;
from openpyxl import load_workbook;
from datetime import *;
from pyomo.environ import *;
from pyomo.environ import *;
from contextlib import suppress;
import pyomo.opt;
import time;
from time import strftime;
import os;
import datetime as dt;
from ctypes import create_unicode_buffer, windll;
import xlwt;
import win32com.client;
import traceback;
import shutil;


class SaidaDados:
    
    def __init__(self, sistema, problema, path, plan_dados, pasta_cod, nomeSubsistemas):
        # recebe como parametro o sistema em que estao as informacoes e o problema com o modelo
        self.sin = sistema;
        self.modelo = problema.modelo;
        self.caminho = path;
        self.planilha = plan_dados;
        self.vMDI = pasta_cod;
        self.nomeSubs = nomeSubsistemas;

        #transformando short path em long
        BUFFER_SIZE = 500;
        buffer = create_unicode_buffer(BUFFER_SIZE);
        get_long_path_name = windll.kernel32.GetLongPathNameW;
        get_long_path_name(str(self.caminho), buffer, BUFFER_SIZE);
        self.caso = buffer.value[-9:];
        self.caso = self.caso.replace("\\","");
        
        # chama os metodos para imprimir as saidas
        self.imprimeTotaisGeracao();
        self.imprimeSaidaExpansao();
        self.imprimeSaidaIntercambio();
        try:    
            self.imprimeResumoExecutivo();
        except:
            print("Erro na impressao da planilha Resumo");
            print(traceback.format_exc());
        self.imprimeBalancoPotenciaSIN();
        self.imprimeCustosPeriodo();
        self.imprimeCustosCenarios();

        return;

    def imprimeLog(self, tempo_inicial):
        self.start = tempo_inicial;
        self.end = time.localtime();
        tempo_total = int(time.mktime(self.end) - time.mktime(self.start));

        # cria o arquivo txt
        saidaResul = open(self.caminho + "log.txt", "w");

        # escreve o usuario que fez executou o caso
        saidaResul.write("Executado por: " + str(os.getlogin()) + "\n");

        #transformando short path em long
        BUFFER_SIZE = 500;
        buffer = create_unicode_buffer(BUFFER_SIZE);
        get_long_path_name = windll.kernel32.GetLongPathNameW;
        get_long_path_name(str(self.planilha), buffer, BUFFER_SIZE);
        vModeloJorge = buffer.value;
        
        # escreve versao do MDI
        saidaResul.write("Versao MDI: " + self.vMDI + "\n");

        # escreve versao da planilha
        saidaResul.write("Versao Planilha: " + vModeloJorge + "\n");

        # escreve hora e data de inicio da execucao
        saidaResul.write("Inicio da Execucao: " + strftime("%d %b %Y %H:%M:%S", self.start) + "\n");

        # escreve hora e data de termino da execucao
        saidaResul.write("Termino da Execucao: " + strftime("%d %b %Y %H:%M:%S", self.end) + "\n");

        # escreve tempo total de execucao
        saidaResul.write("Tempo de Execucao: " + strftime("%H:%M", dt.datetime(2018,1,1,int(tempo_total/3600),int((tempo_total%3600)/60),0).timetuple()) + "\n");

        # fecha o arquivo
        saidaResul.close();
        
        return;

    def imprimeTotaisGeracao(self):
        
        somaTerm = 0;
        m = self.modelo;
        
        # cria o arquivo txt
        saidaResul = open(self.caminho + "balancoEnergiaSIN.txt", "w");
        
        listaTerm = [usina for k,usina in self.sin.listaGeralTerm.items()];
        def getKey(u):
            return u.indexUsinaInterno;
        listaTerm.sort(key=getKey)
        
        # imprime o cabecalho
        saidaResul.write("Cenario;");
        saidaResul.write("Sistema;");
        saidaResul.write("Patamar;");
        saidaResul.write("Periodo;");
        saidaResul.write("Demanda;");
        saidaResul.write("Deficit;");
        saidaResul.write("Hidro_Exist;");
        saidaResul.write("PCH_Exist;");
        saidaResul.write("EOL_Exist;");
        saidaResul.write("UFV_Exist;");
        saidaResul.write("BIO_Exist;");
        for rev_ind in m.projReversivel: 
            saidaResul.write("Prod_" + rev_ind + ";");
            saidaResul.write("Arm_" + rev_ind + ";");
        for renov_ind in m.projRenovCont: 
            saidaResul.write("RenovCont_" + renov_ind + ";");
        for uheNova_ind in m.projUHENova: 
            saidaResul.write("UHENova_" + uheNova_ind + ";");
        for term in listaTerm: 
            saidaResul.write("TermicaExist_" + term.nomeUsina + ";");
        for projterm_ind in m.projTermCont: 
            saidaResul.write("TermCont_" + projterm_ind.replace(";",",") + ";");
        for subsis in m.subsistemas: 
            saidaResul.write("Enviado Sistema" + str(subsis+1) + ";");
        for subsis in m.subsistemas: 
            saidaResul.write("Recebido Sistema" + str(subsis+1) + ";");
               
        # pula uma linha no arquivo
        saidaResul.write("\n");
        
        # imprime totais de cada tipo de geracao por condicao, subsistema e periodo
        for icen in (m.condicoes):
            if (self.sin.tipoCombHidroEol == "completa"):
                icond = icen;
            elif (self.sin.tipoCombHidroEol == "intercalada"):
                icond = icen % self.sin.numEol;
            else:
                print("opcao de combinacao de series hidrologicas com eolicas nao marcada");
            for isis in (m.subsistemas):
                for ipat in (m.patamares):
                    for iper in (m.periodosTotal):
                        
                        saidaResul.write(str(icen+1) + ";");
                        saidaResul.write(str(isis+1) + ";");
                        saidaResul.write(str(ipat+1) + ";");
                        saidaResul.write(str(iper+1) + ";");
                        saidaResul.write(str(m.demanda[isis,ipat,iper,icen]) + ";");
                        saidaResul.write(str(m.deficit[isis,ipat,iper,icen].value) + ";");
                        saidaResul.write(str(m.prodHidroExist[isis,ipat,iper,icen].value) + ";");
                        saidaResul.write(str(self.sin.subsistemas[isis].montanteRenovExPCH[iper] * self.sin.subsistemas[isis].fatorPatPCH[ipat][iper%12]) + ";");
                        saidaResul.write(str(self.sin.subsistemas[isis].montanteRenovExEOL[iper] * self.sin.subsistemas[isis].fatorPatEOLEx[ipat][iper%12]) + ";");
                        saidaResul.write(str(self.sin.subsistemas[isis].montanteRenovExUFV[iper] * self.sin.subsistemas[isis].fatorPatUFVEx[ipat][iper%12]) + ";");
                        saidaResul.write(str(self.sin.subsistemas[isis].montanteRenovExBIO[iper] * self.sin.subsistemas[isis].fatorPatBIO[ipat][iper%12]) + ";");

                        # impressao da geracao e bombeamento das reversiveis
                        for rev in m.projReversivel:
                            v1=0;
                            v2=0;
                            proj = self.sin.listaGeralProjReversivel[rev]; # pega o projeto                        
                            # so imprime se for do submercado em questao
                            if proj.sis_index == (isis+1):
                                v1 = m.prodReversivel[rev, ipat, iper, icen].value;
                                v2 = m.bombReversivel[rev, ipat, iper, icen].value;
                        
                            saidaResul.write(str(v1) + ";"); # computa e imprime a contribuicao
                            saidaResul.write(str(v2) + ";"); # computa e imprime a contribuicao

                        # impressao das geracoes das renovaveis continuas
                        for renov_ind in m.projRenovCont:
                            proj = self.sin.listaGeralProjRenov[renov_ind]; # pega o projeto
                            v=0;
                            # so imprime se for do submercado em questao
                            if proj.sis_index == (isis+1):
                                if (proj.tipo == 'BIO'):
                                    v=(m.capRenovCont[renov_ind,iper].value)*(proj.fatorCapacidade[iper % 12])*(self.sin.subsistemas[isis].fatorPatBIO[ipat][iper%12]);
                                if (proj.tipo == 'UFV'):
                                    v=(m.capRenovCont[renov_ind,iper].value)*(proj.fatorCapacidade[iper % 12])*(self.sin.subsistemas[isis].fatorPatUFV[ipat][iper%12]);
                                if (proj.tipo == 'PCH'):
                                    v=(m.capRenovCont[renov_ind,iper].value)*(proj.fatorCapacidade[iper % 12])*(self.sin.subsistemas[isis].fatorPatPCH[ipat][iper%12]);
                                if (proj.tipo == 'EOF'):
                                    v=(m.capRenovCont[renov_ind,iper].value)*(proj.fatorCapacidade[iper % 12])*(self.sin.subsistemas[isis].fatorPatEOF[ipat][iper%12]);
                                if (proj.tipo == 'EOL'):
                                    v=(m.capRenovCont[renov_ind,iper].value)*(proj.seriesEolicas[icond][iper % 12])*(self.sin.subsistemas[isis].fatorPatEOL[ipat][iper%12]);
                            saidaResul.write(str(v) + ";"); # computa e imprime a contribuicao
                            
                        # impressao das geracoes das uhes novas
                        for uhe_ind in m.projUHENova:
                            v=0;
                            proj = self.sin.listaGeralProjUHE[uhe_ind]; # pega o projeto                        
                            # so imprime se for do submercado em questao
                            if self.sin.listaGeralProjUHE[uhe_ind].sis_index == (isis+1):
                                v = m.prodHidroNova[uhe_ind, ipat, iper, icen].value;
                            saidaResul.write(str(v) + ";"); # computa e imprime a contribuicao
                            
                        # impressao das geracoes termicas existentes
                        for term in listaTerm:
                            v=0;
                            # so imprime valor nao nulo se a termica for do subsistema
                            if term.sis_index == (isis+1):
                                v=m.prodTerm[term.nomeUsina,ipat,iper,icen].value;
                            saidaResul.write(str(v) + ";"); # imprime a contribuicao
                            
                        # impressao das geracoes termicas continuas
                        for projterm_ind in m.projTermCont:
                            v=0;
                            # so imprime se for do submercado em questao
                            if self.sin.listaGeralProjTerm[projterm_ind].sis_index == (isis+1):
                                v=m.prodTermCont[projterm_ind,ipat,iper,icen].value;
                            saidaResul.write(str(v) + ";"); # computa e imprime a contribuicao
                            
                        # impressao dos intercambios enviados
                        for jsis in m.subsistemas: 
                            saidaResul.write(str(m.interc[isis,jsis,ipat,iper,icen].value) + ";");
                        
                        # impressao dos intercambios recebidos
                        for jsis in m.subsistemas:
                            saidaResul.write(str((1-self.sin.subsistemas[isis].perdasInterc[jsis][ipat][iper])*m.interc[jsis,isis,ipat,iper,icen].value) + ";");

                        saidaResul.write("\n");

        # fecha o arquivo
        saidaResul.close();
        
        return;
    
    def imprimeSaidaExpansao(self):        
        
        # cria o arquivo txt
        saidaResul = open(self.caminho + "saidaExpansao.txt", "w");
        
        # imprime cabecalhos              
        for term in self.modelo.projTermCont:
            saidaResul.write(term.replace(";",",") + ";");            
        for renov in self.modelo.projRenovCont:
            saidaResul.write(renov + ";");
        for proj in self.modelo.projReversivel:
            saidaResul.write(proj + ";");           
        saidaResul.write("\n");
        
        # imprime cada uma das linhas
        modelo = self.modelo;
        sin = self.sin;
        
        for iper in modelo.periodosTotal:
            # percorre todos os conjuntos e imprime o resultado de cada um
            for term in modelo.projTermCont:
                saidaResul.write(str(modelo.capTermCont[term,iper].value*modelo.sin.listaGeralProjTerm[term].potUsina) + ";");            
            for renov in modelo.projRenovCont:
                saidaResul.write(str(modelo.capRenovCont[renov,iper].value) + ";");
            for proj in modelo.projReversivel:
                saidaResul.write(str(modelo.capReversivel[proj,iper].value) + ";");
            saidaResul.write("\n");                                         
        
        # fecha o arquivo        
        saidaResul.close();
        
        # abre o arquivo para a saidas
        saidaResul = open(self.caminho + "saidaExpansaoBinaria.txt", "w");
        
        # cabecalho
        saidaResul.write("HIDRO\n");
        
        # percorre todas as hidro
        for hidro in modelo.projUHENova:
            # pega o mes que a usina entrou
            mes_entrada = modelo.sin.numMeses # inicializa no final do horizonte
            for per in modelo.periodosTotal: # percorre todos os periodos
                if (modelo.investHidro[hidro,per] == 1): mes_entrada = per  # no periodo que entrou atribui o periodo
            
            # so imprime se a usina tiver entrado
            if mes_entrada < modelo.sin.numMeses:
                # imprime a entrada
                saidaResul.write(str(int(sin.listaGeralProjUHE[hidro].indexUsinaExterno)) + " no tempo " + str(int(mes_entrada+1)) + "\n");
        
        # fecha o arquivo
        saidaResul.close();

        return;
    
    def imprimeSaidaIntercambio(self):
        
        # cria o arquivo txt
        saidaResul = open(self.caminho + "saidaExpIntercambio.txt", "w");
        
        # percorre cada um dos cenarios e periodos e imprime os intercambios entre os subsistemas
        modelo = self.modelo; 
        
        # cabecalho
        saidaResul.write("Periodo, Origem, Destino, Patamar, CapExistente, ExpansaoInt\n");
        for iper in (self.modelo.periodosTotal):
            for isis in modelo.subsistemas:
                for jsis in modelo.subsistemas:
                    for ipat in modelo.patamares:
                        saidaResul.write(str(iper) + "," + str(isis) + "," + str(jsis) + "," + str(ipat) + "," + str(self.sin.subsistemas[isis].capExistente[jsis][ipat][iper]) + "," + str(modelo.capExpInter[isis,jsis,iper].value) + "\n");
        
        # fecha o arquivo
        saidaResul.close();
        
        return;
    
    def imprimeResumoExecutivo(self):

        # abre o arquivo e pega a aba
        wb = load_workbook(self.caminho + "Resumo.xlsx");  	
        aba = wb.get_sheet_by_name("Resumo Executivo");
        
        # pega objetos basicos
        modelo = self.modelo;
        sin = self.sin;
        
        # pega o ano inicial da impressao
        anoInicial = int(aba.cell(row=7,column=5).value - self.sin.anoInicial);
        ncols = self.sin.numAnos-anoInicial;
        
        # vetor para armazenar totais
        total = [0] * (ncols+1);
                       
        # percorre as linhas para saber o tipo de fonte
        linha = 0;
        while (aba.cell(row=8,column=23).offset(row=linha).value is not None):
            # pega o tipo
            tipoProj = aba.cell(row=8,column=23).offset(row=linha).value;
            
            # pega os projetos
            projs = aba.cell(row=8,column=24).offset(row=linha).value;
            
            # prepara a lista de projetos
            lista_projs = [];
            if (projs != "") and (projs is not None):
                projs = str(projs);
                lista_projs = projs.split(";");
                # converte para inteiro
                lista_projs = [int(float(item)) for item in lista_projs];
            
            # inicializa o valor acumulado
            val_acum = [0];
            val_inv = 0;
            ano_PDE = 10;
            val_inv_PDE = 0;
            
            with suppress(Exception):
                # percorre os anos
                for iano in range(anoInicial,self.sin.numAnos):
                    # pega o iper como dezembro do ano
                    iper = int((iano+1)*12-1);
                    
                    # verifica o tipo de projetos
                    if tipoProj == "Hidro":
                        # pega o valor do periodo de todos os projetos
                        val_acum.append(sum(sin.listaGeralProjUHE[hidro].potUsina*sum(modelo.investHidro[hidro,iperaux].value for iperaux in range(0,iper+1)) for hidro in modelo.projUHENova));
                        val_inv += sum(modelo.custoInvHidro[hidro]*sum(modelo.investHidro[hidro,iperaux].value for iperaux in range(0,iper)) for hidro in modelo.projUHENova for iper_all in range(iano*12, (iano+1)*12));
                        if iano <= ano_PDE:
                            val_inv_PDE += sum(modelo.custoInvHidro[hidro]*sum(modelo.investHidro[hidro,iperaux].value for iperaux in range(0,iper)) for hidro in modelo.projUHENova for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "Reversivel":
                        # pega o valor do periodo de todos os projetos
                        val_acum.append(sum(modelo.capReversivel[sin.listaIndGeralProjReversivel[revers-1].nomeUsina,iper].value for revers in lista_projs));
                        val_inv += sum(modelo.custoInvReversivel[sin.listaIndGeralProjReversivel[revers-1].nomeUsina]*modelo.capReversivel[sin.listaIndGeralProjReversivel[revers-1].nomeUsina,iper_all].value for revers in lista_projs for iper_all in range(iano*12, (iano+1)*12));
                        if iano <= ano_PDE:
                            val_inv_PDE += sum(modelo.custoInvReversivel[sin.listaIndGeralProjReversivel[revers-1].nomeUsina]*modelo.capReversivel[sin.listaIndGeralProjReversivel[revers-1].nomeUsina,iper_all].value for revers in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "RenovCont":
                        # pega o valor do periodo de todos os projetos
                        val_acum.append(sum(modelo.capRenovCont[sin.listaIndGeralProjRenov[renov-1].nomeUsina,iper].value for renov in lista_projs));
                        val_inv += sum(modelo.custoInvRenovCont[sin.listaIndGeralProjRenov[renov-1].nomeUsina]*modelo.capRenovCont[sin.listaIndGeralProjRenov[renov-1].nomeUsina,iper_all].value for renov in lista_projs for iper_all in range(iano*12, (iano+1)*12));
                        if iano <= ano_PDE:
                            val_inv_PDE += sum(modelo.custoInvRenovCont[sin.listaIndGeralProjRenov[renov-1].nomeUsina]*modelo.capRenovCont[sin.listaIndGeralProjRenov[renov-1].nomeUsina,iper_all].value for renov in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "TermCont":
                        # pega o valor do periodo de todos os projetos
                        val_acum.append(sum(modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper].value/sin.listaIndGeralProjTerm[term].fdisp for term in lista_projs));
                        val_inv += sum(modelo.custoInvProjTerm[sin.listaIndGeralProjTerm[term].nomeUsina]*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper_all].value for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));
                        if iano <= ano_PDE:
                            val_inv_PDE += sum(modelo.custoInvProjTerm[sin.listaIndGeralProjTerm[term].nomeUsina]*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper_all].value for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "TermContT":
                        # pega o valor do periodo de todos os projetos
                        val_acum.append(sum(modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper].value/sin.listaIndGeralProjTerm[term].fdisp for term in lista_projs));
                        val_inv += sum(modelo.custoInvProjTerm[sin.listaIndGeralProjTerm[term].nomeUsina]*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper_all].value for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));
                        if iano <= ano_PDE:
                            val_inv_PDE += sum(modelo.custoInvProjTerm[sin.listaIndGeralProjTerm[term].nomeUsina]*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper_all].value for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    if tipoProj == "TermInt":
                        # pega o valor do periodo de todos os projetos
                        val_acum.append(sum(sin.listaIndGeralProjTerm[term].potUsina*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper].value/sin.listaIndGeralProjTerm[term].fdisp for term in lista_projs));
                        val_inv += sum(modelo.custoInvProjTerm[sin.listaIndGeralProjTerm[term].nomeUsina]*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper_all].value for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));
                        if iano <= ano_PDE:
                            val_inv_PDE += sum(modelo.custoInvProjTerm[sin.listaIndGeralProjTerm[term].nomeUsina]*modelo.capTermCont[sin.listaIndGeralProjTerm[term].nomeUsina,iper_all].value for term in lista_projs for iper_all in range(iano*12, (iano+1)*12));

                    # escreve o valor incremental
                    col = iano-anoInicial;
                    val_inc = val_acum[col+1]-val_acum[col];
                    aba.cell(row=8,column=5).offset(row=linha,column=col).value = val_inc;
                    
                    if tipoProj != "TermContT":
                    # soma no total geral
                        total[col] += val_inc;
                        total[ncols] += val_inc;

                # imprime o total da fonte
                aba.cell(row=8,column=5).offset(row=linha,column=col+1).value = val_acum[col+1];

                # imprime o total para ano final do horizonte decenal
                aba.cell(row=8,column=5).offset(row=linha,column=col+2).value = val_acum[9];

                # imprime o total de investimento - divide por 1 milhao por definicao de unidade na tabela do excel
                aba.cell(row=8,column=5).offset(row=linha,column=col+4).value = val_inv/1000000;

                # imprime o total de investimento no ano alvo - divide por 1 milhao por definicao de unidade na tabela do excel
                aba.cell(row=8,column=5).offset(row=linha,column=col+5).value = val_inv_PDE/1000000;
            
            # proxima linha
            linha += 1;
                    
        # imprime os totais gerais anuais
        for col in range(ncols+1): 
            aba.cell(row=8,column=5).offset(row=linha,column=col).value = total[col];
            
        # guarda as informacoes
        (self.nlinhas, self.ncols) = (linha,ncols);

        # imprime a funcao objetivo
        aba.cell(row=self.nlinhas + 16,column=17).value = value(modelo.obj)/1000000;
        
        # imprime a saida de UHE novas
        self.imprimeResumoUHE(aba, self.nlinhas+8);

        # imprime a expansao da transmissao
        self.imprimeResumoTransmissao(aba, self.nlinhas+8);

        # fecha a planilha de resumo executivo
        wb.save(self.caminho + "Resumo.xlsx");
        shutil.copyfile(self.caminho + "Resumo.xlsx", self.caminho + "Resumo" + self.caso + ".xlsx");

        return;
        
    def imprimeResumoUHE(self, aba, linha_inicio):
        m = self.modelo;
        nmeses = self.sin.numMeses;
        pos = linha_inicio;

        # limpa celulas
        for row in aba['D' + str(pos + 8) + ':G' + str(pos + 37)]:
            for cell in row:
                cell.value = None

        # monta um dicionario com os nomes e os periodos
        listaUHE = {h_ind: (nmeses - sum(sum(m.investHidro[h_ind,iperaux].value for iperaux in range(0,per+1))  for per in m.periodosTotal)) for h_ind in m.projUHENova};
        
        # ordena a lista pelo periodo de entrada
        for nome in sorted(listaUHE, key=listaUHE.get):
            # imprime apenas aquelas que entrar no horizonte
            per = listaUHE[nome];
            if per < nmeses:
                # pega o projeto
                proj = self.sin.listaGeralProjUHE[nome];
                aba.cell(row=8,column=4).offset(row=pos).value = nome;
                aba.cell(row=8,column=5).offset(row=pos).value = self.nomeSubs[int(proj.sis_index)-1];
                aba.cell(row=8,column=6).offset(row=pos).value = proj.potUsina;
                aba.cell(row=8,column=7).offset(row=pos).value = str(int(per%12+1)) + "/" + str(int(int(self.sin.anoInicial)+per//12));
                pos += 1
                
        return;

    def imprimeResumoTransmissao(self, aba, linha_inicio):
        m = self.modelo;
        nmeses = self.sin.numMeses;
        pos = linha_inicio;

        # limpa celulas
        for row in aba['I' + str(pos + 8) + ':L' + str(pos + 60)]:
            for cell in row:
                cell.value = None

        for iper in range(1,nmeses-1):
            for isis in m.subsistemas:
                for jsis in m.subsistemas:
                    if m.capExpInter[isis,jsis,iper].value > m.capExpInter[isis,jsis,iper-1].value:
                        aba.cell(row=8,column=9).offset(row=pos).value = self.nomeSubs[isis];
                        aba.cell(row=8,column=10).offset(row=pos).value = self.nomeSubs[jsis];
                        aba.cell(row=8,column=11).offset(row=pos).value = m.capExpInter[isis,jsis,iper].value - m.capExpInter[isis,jsis,iper-1].value;
                        aba.cell(row=8,column=12).offset(row=pos).value = str(int(iper%12+1)) + "/" + str(int(int(self.sin.anoInicial)+iper//12));
                        pos += 1
                
        return;
    
    def imprimeBalancoPotenciaSIN(self):
        
        somaTerm = 0;
        m = self.modelo;
        s = self.sin;
        
        # cria o arquivo txt
        saidaResul = open(self.caminho + "balancoPotenciaSIN.txt", "w");
        
        # imprime o cabecalho
        saidaResul.write("Cenario;");
        saidaResul.write("Sistema;");
        saidaResul.write("Periodo;");
        saidaResul.write("Demanda * " + str(1 + self.sin.restPot) + ";");
        saidaResul.write("DeficitPot;");
        saidaResul.write("Hidro_Exist;");
        saidaResul.write("PCH_Exist;");
        saidaResul.write("EOL_Exist;");
        saidaResul.write("UFV_Exist;");
        saidaResul.write("BIO_Exist;");
        saidaResul.write("Term_Exist;");
        saidaResul.write("Prod_Armazenamento;");
        for uheNova_ind in m.projUHENova: 
            saidaResul.write("UHENova_" + uheNova_ind + ";");
        saidaResul.write("EOL_Nova;");
        saidaResul.write("UFV_Nova;");
        saidaResul.write("BIO_Nova;");
        saidaResul.write("PCH_Nova;");
        saidaResul.write("Term_Nova;");
        saidaResul.write("Enviado Sistema;");
        saidaResul.write("Recebido Sistema;");

        # pula uma linha no arquivo
        saidaResul.write("\n");
    
        # percorre todos os cenarios, sistemas e periodos
        for icen in (m.condicoes):
            if (self.sin.tipoCombHidroEol == "completa"):
                icond = icen;
            elif (self.sin.tipoCombHidroEol == "intercalada"):
                icond = icen % self.sin.numEol;
            else:
                print("opcao de combinacao de series hidrologicas com eolicas nao marcada");
            for isis in (m.subsistemas):
                for iper in (m.periodosTotal):

                    saidaResul.write(str(icen+1) + ";");
                    saidaResul.write(str(isis+1) + ";");
                    saidaResul.write(str(iper+1) + ";");
                    saidaResul.write(str(m.demanda[isis,0,iper,icen]*(1 + self.sin.restPot)) + ";");
                    saidaResul.write(str(m.deficitPot[isis,iper,icen].value) + ";");
                    saidaResul.write(str(m.prodHidroExist[isis,0,iper,icen].value) + ";");
                    saidaResul.write(str(m.enPCHEx[isis,iper,icen] * s.subsistemas[isis].fatorPatPCH[0][iper%12]) + ";");
                    saidaResul.write(str(m.enEOLEx[isis,iper,icen] * s.subsistemas[isis].fatorPatEOLEx[0][iper%12]) + ";");
                    saidaResul.write(str(m.enUFVEx[isis,iper,icen] * s.subsistemas[isis].fatorPatUFVEx[0][iper%12]) + ";");
                    saidaResul.write(str(m.enBIOEx[isis,iper,icen] * s.subsistemas[isis].fatorPatBIO[0][iper%12]) + ";");
                    saidaResul.write(str(sum(term.potUsina[iper] for term in s.subsistemas[isis].listaTermica)) + ";");
                    saidaResul.write(str(sum(m.prodReversivel[proj.nomeUsina,0,iper,icen].value for proj in s.subsistemas[isis].listaProjReversivel)) + ";");
                    
                    # impressao das geracoes das uhes novas
                    for uhe_ind in m.projUHENova:
                        v=0;
                        proj = self.sin.listaGeralProjUHE[uhe_ind]; # pega o projeto                        
                        # so imprime se for do submercado em questao
                        if self.sin.listaGeralProjUHE[uhe_ind].sis_index == (isis+1):
                            v = m.prodHidroNova[uhe_ind, 0, iper, icen].value;
                        saidaResul.write(str(v) + ";"); # computa e imprime a contribuicao
                
                    saidaResul.write(str(sum(m.capRenovCont[proj.nomeUsina,iper].value*proj.seriesEolicas[icond][iper % 12]*s.subsistemas[isis].fatorPatEOL[0][iper%12] for proj in s.subsistemas[isis].listaProjEOL) + \
                    sum(m.capRenovCont[proj.nomeUsina,iper].value*proj.fatorCapacidade[iper % 12]*s.subsistemas[isis].fatorPatEOF[0][iper%12] for proj in s.subsistemas[isis].listaProjEOF)) + ";");
                    saidaResul.write(str(sum(m.capRenovCont[proj.nomeUsina,iper].value*proj.fatorCapacidade[iper % 12]*s.subsistemas[isis].fatorPatUFV[0][iper%12] for proj in s.subsistemas[isis].listaProjUFV)) + ";");
                    saidaResul.write(str(sum(m.capRenovCont[proj.nomeUsina,iper].value*proj.fatorCapacidade[iper % 12]*s.subsistemas[isis].fatorPatBIO[0][iper%12] for proj in s.subsistemas[isis].listaProjBIO)) + ";");
                    saidaResul.write(str(sum(m.capRenovCont[proj.nomeUsina,iper].value*proj.fatorCapacidade[iper % 12]*s.subsistemas[isis].fatorPatPCH[0][iper%12] for proj in s.subsistemas[isis].listaProjPCH)) + ";");
                    saidaResul.write(str(sum(m.capTermCont[proj.nomeUsina,iper].value*proj.potUsina for proj in s.subsistemas[isis].listaProjTermica)) + ";");
                    saidaResul.write(str(sum(m.intercPot[isis,jsis,iper,icen].value for jsis in m.subsistemas)) + ";");
                    saidaResul.write(str(sum((1-self.sin.subsistemas[isis].perdasInterc[jsis][0][iper])*m.intercPot[jsis,isis,iper,icen].value for jsis in m.subsistemas)) + ";");
                    
                    saidaResul.write("\n");

        # fecha o arquivo
        saidaResul.close();

        return;
    
    def limparDuais(self):
        try:
            # abre o arquivo e pega a aba
            wb = load_workbook(self.caminho + "Resumo.xlsx");        
            aba = wb.get_sheet_by_name("Resumo Executivo");
            
            # pega o ano inicial
            anoInicial = int(aba.cell(row=7,column=5).value - self.sin.anoInicial);

            # limpa as linhas existentes com os duais para quando eles nao forem impressos
            col = 0;
            for iAno in range(anoInicial, self.sin.numAnos):
                aba.cell(row=8,column=5).offset(row=self.nlinhas+2,column=col).value = None; 
                aba.cell(row=8,column=5).offset(row=self.nlinhas+3,column=col).value = None;  
                aba.cell(row=8,column=5).offset(row=self.nlinhas+4,column=col).value = None; 
                col = col+1;  
                
            # fecha a planilha de resumo executivo
            wb.save(self.caminho + "Resumo.xlsx");
            shutil.copyfile(self.caminho + "Resumo.xlsx", self.caminho + "Resumo" + self.caso + ".xlsx");
        except:
            print("Erro ao tentar acessar planilha Resumo");
        return;
    
    def imprimeDuais(self, tipo_dual):
        m = self.modelo;
        
        # pega todos os duais e, pelo tipo, diz em qual linha devem ser escritos
        linha_dual=0;
        den=8760
        if (tipo_dual == "D"):
            duais = {int(float(indAno)) : m.dual.get(m.DualD[indAno]) for indAno in m.anos};
            linha_dual=2;
            
            #imprime txt dos duais
            saidaResul = open(self.caminho + "dualDuplo.txt", "w");
        
            # cabecalho
            saidaResul.write("Ano; Dual\n");
        
            # percorre todos os anos
            for iano in range(1,self.sin.numAnos):             
                # imprime o dual
                dual = ((m.dual.get(m.DualD[iano-1]) - m.dual.get(m.DualD[iano]))/den)*(1+m.sin.taxaDesc)**(iano*12+6);
                saidaResul.write(str(m.sin.anoInicial + iano) + "; " + '{0:.2f}'.format(dual) + "\n");
            # fecha o arquivo
            saidaResul.close();

        elif (tipo_dual == "P"):
            duais = {int(float(indAno)) : m.dual.get(m.DualP[indAno]) for indAno in m.anos};
            linha_dual=1;
            den=1000 # no caso de potencia nao usa o numero de horas so a coversao para kW

            #imprime txt dos duais
            saidaResul = open(self.caminho + "dualPot.txt", "w");
        
            # cabecalho
            saidaResul.write("Ano; Dual\n");
        
            # percorre todos os anos
            for iano in range(1,self.sin.numAnos):             
                # imprime o dual
                dual = ((m.dual.get(m.DualP[iano-1]) - m.dual.get(m.DualP[iano]))/den)*(1+m.sin.taxaDesc)**(iano*12+6);
                saidaResul.write(str(m.sin.anoInicial + iano) + "; " + '{0:.2f}'.format(dual) + "\n");
            # fecha o arquivo
            saidaResul.close();
        elif (tipo_dual == "E"):
            duais = {int(float(indAno)) : m.dual.get(m.DualE[indAno]) for indAno in m.anos};

            #imprime txt dos duais
            saidaResul = open(self.caminho + "dualEnergia.txt", "w");
        
            # cabecalho
            saidaResul.write("Ano; Dual\n");
        
            # percorre todos os anos
            for iano in range(1,self.sin.numAnos):             
                # imprime o dual
                dual = ((m.dual.get(m.DualE[iano-1]) - m.dual.get(m.DualE[iano]))/den)*(1+m.sin.taxaDesc)**(iano*12+6);
                saidaResul.write(str(m.sin.anoInicial + iano) + "; " + '{0:.2f}'.format(dual) + "\n");
            # fecha o arquivo
            saidaResul.close();

        try:
            # abre o arquivo e pega a aba
            wb = load_workbook(self.caminho + "Resumo.xlsx");
            aba = wb.get_sheet_by_name("Resumo Executivo");
        
            # pega o ano inicial
            anoInicial = int(aba.cell(row=7,column=5).value - self.sin.anoInicial);
                
            # percorre as colunas imprimindo os duais escolhidos pelo usuario (CME)
            col = 0;
            for iAno in range(anoInicial, self.sin.numAnos):
                aba.cell(row=8,column=5).offset(row=self.nlinhas+2+linha_dual,column=col).value = ((duais[iAno-1] - duais[iAno])/den)*(1+m.sin.taxaDesc)**(iAno*12+6);
                col = col+1;  
            
            # fecha a planilha de resumo executivo
            wb.save(self.caminho + "Resumo.xlsx");
            shutil.copyfile(self.caminho + "Resumo.xlsx", self.caminho + "Resumo" + self.caso + ".xlsx");
        except:
            print("Erro ao acessar planilha Resumo. DUAL NÃO SERÁ IMPRESSO NA PLLANILHA RESUMO!");
        return;
        
    def imprimeSeriesHidro(self):
        sin = self.sin;
        
        listaUHE = [usina for subsis in sin.subsistemas for usina in subsis.listaUHE];
        listaUHE.extend([usina for subsis in sin.subsistemas for usina in subsis.listaProjUHE]);
        
        # percorre os cenarios
        for icen in range(sin.numHidros):
            # abre o arquivro
            saidaResul = open(self.caminho + "serieHidro" + str(icen) + ".txt", "w");
            
            # percorre primeiramente os projetos
            for proj in listaUHE:
                # imprime o nome da usina
                saidaResul.write(proj.nomeUsina);
                
                # percorre os periodos
                for iper in range(sin.numMeses): 
                    saidaResul.write("," + str(proj.serieHidrologica[icen][iper]));
                
                # proxima linha
                saidaResul.write("\n");
            
            # fecha o arquivo
            saidaResul.close();
        
        return;
     
    def imprimeCustosPeriodo(self):
        modelo = self.modelo;

        # abre o arquivo para a saidas
        saidaResul = open(self.caminho + "custosPeriodos.txt", "w");
        
        # cabecalho
        saidaResul.write("Periodo; Custo Total; Custo GT; Custo Deficit Energia; Custo Deficit Capacidade; Custo Bombeamento; Custo Intercambio; Custo Penalidades GH; Custo Investimento\n");
        
        # percorre todos os periodos
        for iper in modelo.periodosTotal:
            # inicializa a soma do periodos
            custo_tot = modelo.sin.horasMes*(
                        sum(modelo.sin.probHidro[icen]*modelo.prodTerm[term, ipat, iper, icen].value*modelo.cvuTermExist[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.termExist for ipat in modelo.patamares  for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*modelo.prodTermCont[term, ipat, iper, icen].value*modelo.cvuProjTerm[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.projTermCont for ipat in modelo.patamares for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*modelo.deficit[isis, ipat, iper , icen].value*modelo.custoDefc[isis, ipat]*modelo.sin.duracaoPatamar[ipat][iper] for isis in modelo.subsistemas for ipat in modelo.patamares for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*modelo.bombReversivel[rev, ipat, iper , icen].value*modelo.sin.pldMin*modelo.sin.duracaoPatamar[ipat][iper] for rev in modelo.projReversivel for ipat in modelo.patamares for icen in modelo.condicoes)) + \
                        sum(modelo.sin.probHidro[icen]*0.0005*modelo.interc[isis,jsis,ipat,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas for ipat in modelo.patamares for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*0.0005*modelo.intercPot[isis,jsis,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*modelo.penalidadeGHMinExist[isis,ipat,iper,icen].value for isis in modelo.subsistemas for ipat in modelo.patamares for icen in modelo.condicoes)*9999 + \
                        sum(modelo.sin.probHidro[icen]*modelo.penalidadeGHMinNova[iuhe,ipat,iper,icen].value for iuhe in modelo.projUHENova for ipat in modelo.patamares for icen in modelo.condicoes)*9999 + \
                        sum(modelo.sin.probHidro[icen]*modelo.deficitPot[isis, iper, icen].value*modelo.sin.custoDefPot for isis in modelo.subsistemas for icen in modelo.condicoes) + \
                        sum(modelo.custoInvHidro[hidro] * sum(modelo.investHidro[hidro,tau].value for tau in range(iper+1)) for hidro in modelo.projUHENova ) + \
                        sum(modelo.custoInvRenovCont[renov] * modelo.capRenovCont[renov,iper].value for renov in modelo.projRenovCont ) + \
                        sum(1000*modelo.sin.subsistemas[isis].custoExpansao[jsis] * modelo.capExpInter[isis,jsis,iper].value for isis in modelo.subsistemas for jsis in range(isis, modelo.sin.nsis) ) + \
                        sum(modelo.custoInvProjTerm[term] * modelo.capTermCont[term,iper].value / modelo.fdispProjTerm[term] for term in modelo.projTermCont ) + \
                        sum(modelo.custoInvReversivel[iproj] * modelo.capReversivel[iproj,iper].value for iproj in modelo.projReversivel);

            custo_gt = modelo.sin.horasMes*(
                        sum(modelo.sin.probHidro[icen]*modelo.prodTerm[term, ipat, iper, icen].value*modelo.cvuTermExist[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.termExist for ipat in modelo.patamares  for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*modelo.prodTermCont[term, ipat, iper, icen].value*modelo.cvuProjTerm[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.projTermCont for ipat in modelo.patamares for icen in modelo.condicoes));

            custo_de = modelo.sin.horasMes*(sum(modelo.sin.probHidro[icen]*modelo.deficit[isis, ipat, iper , icen].value*modelo.custoDefc[isis, ipat]*modelo.sin.duracaoPatamar[ipat][iper] for isis in modelo.subsistemas for ipat in modelo.patamares for icen in modelo.condicoes));

            custo_dp =  sum(modelo.sin.probHidro[icen]*modelo.deficitPot[isis, iper, icen].value*modelo.sin.custoDefPot for isis in modelo.subsistemas for icen in modelo.condicoes);

            custo_br = modelo.sin.horasMes*(sum(modelo.sin.probHidro[icen]*modelo.bombReversivel[rev, ipat, iper , icen].value*modelo.sin.pldMin*modelo.sin.duracaoPatamar[ipat][iper] for rev in modelo.projReversivel for ipat in modelo.patamares for icen in modelo.condicoes));

            custo_int = sum(modelo.sin.probHidro[icen]*0.0005*modelo.interc[isis,jsis,ipat,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas for ipat in modelo.patamares for icen in modelo.condicoes) + \
                        sum(modelo.sin.probHidro[icen]*0.0005*modelo.intercPot[isis,jsis,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas for icen in modelo.condicoes);

            custo_pen = sum(modelo.sin.probHidro[icen]*modelo.penalidadeGHMinExist[isis,ipat,iper,icen].value for isis in modelo.subsistemas for ipat in modelo.patamares for icen in modelo.condicoes)*9999 + \
                        sum(modelo.sin.probHidro[icen]*modelo.penalidadeGHMinNova[iuhe,ipat,iper,icen].value for iuhe in modelo.projUHENova for ipat in modelo.patamares for icen in modelo.condicoes)*9999;

            custo_inv = sum(modelo.custoInvHidro[hidro] * sum(modelo.investHidro[hidro,tau].value for tau in range(iper+1)) for hidro in modelo.projUHENova ) + \
                        sum(modelo.custoInvRenovCont[renov] * modelo.capRenovCont[renov,iper].value for renov in modelo.projRenovCont ) + \
                        sum(1000*modelo.sin.subsistemas[isis].custoExpansao[jsis] * modelo.capExpInter[isis,jsis,iper].value for isis in modelo.subsistemas for jsis in range(isis, modelo.sin.nsis) ) + \
                        sum(modelo.custoInvProjTerm[term] * modelo.capTermCont[term,iper].value / modelo.fdispProjTerm[term] for term in modelo.projTermCont ) + \
                        sum(modelo.custoInvReversivel[iproj] * modelo.capReversivel[iproj,iper].value for iproj in modelo.projReversivel);
                        
            # imprime o custo
            saidaResul.write(str(iper + 1) + "; " + str(custo_tot) + "; " + str(custo_gt) + "; " + str(custo_de) + "; " + str(custo_dp) + "; " + str(custo_br) + "; " + str(custo_int) + "; " + str(custo_pen) + "; " + str(custo_inv) + "\n");
            
        # fecha o arquivo
        saidaResul.close();

    def imprimeCustosCenarios(self):
        modelo = self.modelo;

        # abre o arquivo para a saidas
        saidaResul = open(self.caminho + "custosCenarios.txt", "w");
        
        # cabecalho
        saidaResul.write("Cenario; Periodo; Custo Total; Custo GT; Custo Deficit Energia; Custo Deficit Capacidade; Custo Bombeamento; Custo Intercambio; Custo Penalidades GH\n");
        
        # percorre todos os cenarios
        for icen in modelo.condicoes:
            # percorre todos os periodos
            for iper in modelo.periodosTotal:
                # inicializa a soma do periodos
                custo_tot = modelo.sin.horasMes*(
                            sum(modelo.prodTerm[term, ipat, iper, icen].value*modelo.cvuTermExist[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.termExist for ipat in modelo.patamares ) + \
                            sum(modelo.prodTermCont[term, ipat, iper, icen].value*modelo.cvuProjTerm[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.projTermCont for ipat in modelo.patamares) + \
                            sum(modelo.deficit[isis, ipat, iper , icen].value*modelo.custoDefc[isis, ipat]*modelo.sin.duracaoPatamar[ipat][iper] for isis in modelo.subsistemas for ipat in modelo.patamares) + \
                            sum(modelo.bombReversivel[rev, ipat, iper , icen].value*modelo.sin.pldMin*modelo.sin.duracaoPatamar[ipat][iper] for rev in modelo.projReversivel for ipat in modelo.patamares)) + \
                            sum(0.0005*modelo.interc[isis,jsis,ipat,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas for ipat in modelo.patamares) + \
                            sum(0.0005*modelo.intercPot[isis,jsis,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas) + \
                            sum(modelo.penalidadeGHMinExist[isis,ipat,iper,icen].value for isis in modelo.subsistemas for ipat in modelo.patamares)*9999 + \
                            sum(modelo.penalidadeGHMinNova[iuhe,ipat,iper,icen].value for iuhe in modelo.projUHENova for ipat in modelo.patamares)*9999 + \
                            sum(modelo.deficitPot[isis, iper, icen].value*modelo.sin.custoDefPot for isis in modelo.subsistemas) + \
                            sum(modelo.custoInvHidro[hidro] * sum(modelo.investHidro[hidro,tau].value for tau in range(iper+1)) for hidro in modelo.projUHENova ) + \
                            sum(modelo.custoInvRenovCont[renov] * modelo.capRenovCont[renov,iper].value for renov in modelo.projRenovCont ) + \
                            sum(1000*modelo.sin.subsistemas[isis].custoExpansao[jsis] * modelo.capExpInter[isis,jsis,iper].value for isis in modelo.subsistemas for jsis in range(isis, modelo.sin.nsis) ) + \
                            sum(modelo.custoInvProjTerm[term] * modelo.capTermCont[term,iper].value / modelo.fdispProjTerm[term] for term in modelo.projTermCont ) + \
                            sum(modelo.custoInvReversivel[iproj] * modelo.capReversivel[iproj,iper].value for iproj in modelo.projReversivel);

                custo_gt = modelo.sin.horasMes*(
                            sum(modelo.prodTerm[term, ipat, iper, icen].value*modelo.cvuTermExist[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.termExist for ipat in modelo.patamares ) + \
                            sum(modelo.prodTermCont[term, ipat, iper, icen].value*modelo.cvuProjTerm[term, iper]*modelo.sin.duracaoPatamar[ipat][iper] for term in modelo.projTermCont for ipat in modelo.patamares));

                custo_de = modelo.sin.horasMes*(sum(modelo.deficit[isis, ipat, iper , icen].value*modelo.custoDefc[isis, ipat]*modelo.sin.duracaoPatamar[ipat][iper] for isis in modelo.subsistemas for ipat in modelo.patamares));

                custo_dp =  sum(modelo.deficitPot[isis, iper, icen].value*modelo.sin.custoDefPot for isis in modelo.subsistemas);

                custo_br = modelo.sin.horasMes*(sum(modelo.bombReversivel[rev, ipat, iper , icen].value*modelo.sin.pldMin*modelo.sin.duracaoPatamar[ipat][iper] for rev in modelo.projReversivel for ipat in modelo.patamares));

                custo_int = sum(0.0005*modelo.interc[isis,jsis,ipat,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas for ipat in modelo.patamares) + \
                            sum(0.0005*modelo.intercPot[isis,jsis,iper,icen].value for isis in modelo.subsistemas for jsis in modelo.subsistemas);

                custo_pen = sum(modelo.penalidadeGHMinExist[isis,ipat,iper,icen].value for isis in modelo.subsistemas for ipat in modelo.patamares)*9999 + \
                            sum(modelo.penalidadeGHMinNova[iuhe,ipat,iper,icen].value for iuhe in modelo.projUHENova for ipat in modelo.patamares)*9999;
                            
                # imprime o custo
                saidaResul.write(str(icen + 1) + "; " + str(iper + 1) + "; " + str(custo_tot) + "; " + str(custo_gt) + "; " + str(custo_de) + "; " + str(custo_dp) + "; " + str(custo_br) + "; " + str(custo_int) + "; " + str(custo_pen) + "\n");
            
        # fecha o arquivo
        saidaResul.close();